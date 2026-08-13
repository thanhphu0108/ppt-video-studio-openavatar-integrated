from __future__ import annotations

import argparse
import hashlib
import json
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.settings import get_settings
from services.errors import VoiceCloneServiceError
from services.synthesis_service import SynthesisService


def _header_key(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return _header_key(value) in {"true", "1", "yes", "y", "x", "có", "co", "xuất", "xuat"}


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return max(0.0, float(value))
    except (TypeError, ValueError):
        return default


def _find_column(headers: dict[str, int], *candidates: str, required: bool = False) -> int | None:
    for candidate in candidates:
        index = headers.get(_header_key(candidate))
        if index is not None:
            return index
    if required:
        raise ValueError(f"Thiếu cột bắt buộc: {candidates[0]}")
    return None


def read_storyboard(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        if "Storyboard" not in workbook.sheetnames:
            raise ValueError("Không tìm thấy sheet 'Storyboard'.")
        sheet = workbook["Storyboard"]
        rows = sheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if not first_row:
            raise ValueError("Sheet Storyboard không có header.")
        headers = {_header_key(value): index for index, value in enumerate(first_row) if _header_key(value)}
        slide_column = _find_column(headers, "Slide", required=True)
        narration_column = _find_column(headers, "Lời thuyết minh", "Loi thuyet minh", required=True)
        export_column = _find_column(headers, "Xuất", "Xuat", "Export")
        pause_column = _find_column(headers, "Nghỉ sau (giây)", "Nghi sau (giay)", "Pause after (seconds)")

        slides: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            slide_value = row[slide_column] if slide_column < len(row) else None
            if slide_value in {None, ""}:
                continue
            try:
                slide = int(slide_value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Dòng {row_number}: Slide phải là số nguyên.") from exc
            raw_text = row[narration_column] if narration_column < len(row) else ""
            text = str(raw_text or "").strip()
            enabled = True if export_column is None else _as_bool(row[export_column] if export_column < len(row) else False)
            pause_after = _as_float(row[pause_column] if pause_column is not None and pause_column < len(row) else 0.0)
            if enabled:
                if not text:
                    raise ValueError(f"Dòng {row_number}, slide {slide}: 'Lời thuyết minh' đang trống.")
                slides.append({"slide": slide, "text": text, "pause_after": pause_after})
        if not slides:
            raise ValueError("Không có slide nào có cột 'Xuất' = TRUE.")
        if len({item["slide"] for item in slides}) != len(slides):
            raise ValueError("Storyboard có số Slide trùng nhau.")
        return slides
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Tổng hợp audio local từ sheet Storyboard trong Excel.")
    parser.add_argument("storyboard", help="Đường dẫn storyboard.xlsx")
    parser.add_argument("--voice-id", default=None, help="Voice profile trong config/voices.json (mặc định: default)")
    parser.add_argument("--model", default=None, help="f5-tts (mặc định) hoặc dummy cho kiểm thử")
    parser.add_argument("--reference-audio", default=None, help="Override audio mẫu, ví dụ voices/default/reference.wav")
    parser.add_argument("--reference-text", default=None, help="Transcript chính xác của audio mẫu")
    parser.add_argument(
        "--confirm-voice-use",
        action="store_true",
        help="Bắt buộc khi dùng --reference-audio: xác nhận bạn có quyền sử dụng giọng mẫu.",
    )
    parser.add_argument("--format", choices=("wav", "mp3"), default="wav", help="Mặc định WAV cho Wav2Lip")
    parser.add_argument("--speed", type=float, default=1.0)
    parser.add_argument("--output-dir", default=None, help="Mặc định generated_audio/")
    parser.add_argument("--continue-on-error", action="store_true", help="Tiếp tục các slide sau khi một slide lỗi")
    return parser


def run(args: argparse.Namespace) -> int:
    try:
        slides = read_storyboard(args.storyboard)
    except (OSError, ValueError) as exc:
        print(f"Lỗi storyboard: {exc}", file=sys.stderr)
        return 2
    if args.reference_audio and not args.confirm_voice_use:
        print("Cần thêm --confirm-voice-use khi dùng --reference-audio.", file=sys.stderr)
        return 2

    settings = get_settings()
    service = SynthesisService(settings)
    service.warm_up()
    output_dir = Path(args.output_dir).resolve() if args.output_dir else settings.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    run_id = str(uuid.uuid4())
    manifest: dict[str, Any] = {
        "run_id": run_id,
        "voice_id": args.voice_id or ("uploaded" if args.reference_audio else settings.default_voice_id),
        "model": args.model or settings.engine,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "slides": [],
    }
    total = len(slides)
    had_error = False
    for index, item in enumerate(slides, start=1):
        started = time.perf_counter()
        slide = item["slide"]
        try:
            rendered = service.synthesize(
                model=args.model,
                voice_id=args.voice_id,
                reference_audio=args.reference_audio,
                reference_text=args.reference_text,
                text=item["text"],
                speed=args.speed,
                output_format=args.format,
                output_dir=output_dir,
                output_name=f"slide_{slide:03d}",
            )
            manifest["slides"].append(
                {
                    "slide": slide,
                    "text_hash": hashlib.sha256(rendered.normalized_text.encode("utf-8")).hexdigest(),
                    "audio": rendered.audio_path.name,
                    "duration": round(rendered.duration_seconds, 3),
                    "pause_after": item["pause_after"],
                    "status": "SUCCESS",
                    "warnings": rendered.warnings,
                    "cache_hit": rendered.cache_hit,
                }
            )
            print(f"[{index}/{total}] Slide {slide} ... done {rendered.duration_seconds:.1f}s")
        except VoiceCloneServiceError as exc:
            had_error = True
            manifest["slides"].append(
                {"slide": slide, "pause_after": item["pause_after"], "status": "FAILED", "error_code": exc.code, "message": exc.message}
            )
            print(f"[{index}/{total}] Slide {slide} ... failed {exc.code}: {exc.message}", file=sys.stderr)
            if not args.continue_on_error:
                break
        except Exception as exc:  # defensive so a manifest still explains a batch failure
            had_error = True
            manifest["slides"].append({"slide": slide, "pause_after": item["pause_after"], "status": "FAILED", "error_code": "UNEXPECTED_ERROR", "message": str(exc)})
            print(f"[{index}/{total}] Slide {slide} ... failed: {exc}", file=sys.stderr)
            if not args.continue_on_error:
                break
    manifest["completed_at"] = datetime.now(timezone.utc).isoformat()
    manifest["success"] = not had_error
    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Manifest: {output_dir / 'manifest.json'}")
    return 1 if had_error else 0


if __name__ == "__main__":
    raise SystemExit(run(build_parser().parse_args()))
